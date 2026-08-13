#!/usr/bin/env python3
"""Regenerate the third-party-producer fixtures.

The corpus already has files written by Excel (via Apache POI's test data) and
by LibreOffice. Those two agree with each other about the one thing an importer
is most tempted to trust: an `.xlsx` stores the writing application's *answer*
next to each formula, and both of them put a correct one there.

Neither of these does, and that is why they are here:

  * **openpyxl** writes `<f>B2*C2</f>` with **no `<v>` at all**. A cell that has
    a formula and no cached value is not a corner case — every file openpyxl has
    ever written looks like this — and an importer that reads values and only
    recalculates on demand shows a blank column.
  * **XlsxWriter** writes `<f>B2*C2</f><v>0</v>`. The cached value is present,
    well-formed, and **wrong**. An importer that trusts the cache shows zeros,
    which is worse than blanks because it looks like an answer.

Both write the same spreadsheet as `libreoffice.sh`, so LibreOffice's computed
values are the oracle for all three: same questions, one implementation's
answers, three different ways of writing it down.

Run it with the libraries in a throwaway environment — they are not a dependency
of anything, and the fixtures they produce are committed:

    python3 -m venv /tmp/v && /tmp/v/bin/pip install xlsxwriter openpyxl
    /tmp/v/bin/python fixtures/tools/producers.py

Output is not byte-stable across library versions, so regenerating changes the
manifest checksums. Expected — update them deliberately, and notice if the
*shapes* moved rather than only the bytes, because the shapes are the point.
"""

import pathlib
import sys

OUT = pathlib.Path("fixtures/corpus")

# The same sheet `libreoffice.sh` converts, with the same expected answers:
# 13.5, 11.88, 840, and 865.38. Kept literal in both places rather than shared
# through a file, because a fixture generator that reads its input from
# somewhere else is one more thing that can silently change.
ROWS = [
    ("Item", "Qty", "Price", None, "Note"),
    ("Widget", 3, 4.5, "=B2*C2", "plain text"),
    ("Gadget", 12, 0.99, "=B3*C3", "comma, inside"),
    ("Gizmo", 7, 120, "=B4*C4", "café ünïcode"),
    (None, None, None, "=SUM(D2:D4)", None),
]


def write_xlsxwriter(path):
    import xlsxwriter

    book = xlsxwriter.Workbook(str(path))
    sheet = book.add_worksheet("Sheet1")
    for r, row in enumerate(ROWS):
        for c, value in enumerate(row):
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                # No `value=` argument, deliberately: that is the default every
                # caller gets, and it stores a cached zero.
                sheet.write_formula(r, c, value)
            else:
                sheet.write(r, c, value)
    book.close()


def write_openpyxl(path):
    import openpyxl

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Sheet1"
    for r, row in enumerate(ROWS, start=1):
        for c, value in enumerate(row, start=1):
            if value is not None:
                sheet.cell(row=r, column=c, value=value)
    book.save(str(path))


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    written = []
    for name, write in (
        ("xlsxwriter-formulas.xlsx", write_xlsxwriter),
        ("openpyxl-formulas.xlsx", write_openpyxl),
    ):
        path = OUT / name
        write(path)
        written.append(path)
        print(f"wrote {path}")

    # Assert the quirk each file exists for is actually in it. A library that
    # changes its default would otherwise leave a fixture that still opens,
    # still passes, and no longer tests anything — which is how a corpus rots
    # without anybody noticing.
    import re
    import zipfile

    for path in written:
        with zipfile.ZipFile(path) as z:
            name = next(n for n in z.namelist() if n.endswith("sheet1.xml"))
            xml = z.read(name).decode("utf-8")
        formulas = re.findall(r"<c[^>]*>(?:(?!</c>).)*?<f[^>]*>.*?</f>.*?</c>", xml)
        if not formulas:
            sys.exit(f"{path}: no formula cells found at all")
        cached = [f for f in formulas if "<v>" in f]
        if "openpyxl" in path.name and cached:
            sys.exit(f"{path}: expected no cached values, found {len(cached)}")
        if "xlsxwriter" in path.name and not any("<v>0</v>" in f for f in cached):
            sys.exit(f"{path}: expected a cached zero, found {cached}")
        print(f"  {path.name}: {len(formulas)} formula cells, {len(cached)} cached")


if __name__ == "__main__":
    main()
